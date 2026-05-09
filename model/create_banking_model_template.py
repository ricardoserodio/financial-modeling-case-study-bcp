"""
Create Banking Model Template – Millennium bcp Case Study

This script creates an Excel template for the simplified banking financial model.

The workbook is created for educational and portfolio purposes only and does not
constitute investment advice.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "banking_model.xlsx"

PERIODS = ["2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]


def style_header(row):
    """Apply basic styling to a header row."""
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(ws):
    """Apply a simple column width adjustment."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3


def create_cover_sheet(wb):
    ws = wb.active
    ws.title = "Cover"

    rows = [
        ["Financial Modeling Case Study – Millennium bcp / Portuguese Listed Bank"],
        [""],
        ["Purpose", "Educational financial modeling case study"],
        ["Company", "Millennium bcp / Banco Comercial Português"],
        ["Model Period", "2022A–2028E"],
        ["Author", "Ricardo Serôdio"],
        [""],
        [
            "Disclaimer",
            "This model is for educational and portfolio purposes only and does not constitute investment advice.",
        ],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    autofit_columns(ws)


def create_sources_sheet(wb):
    ws = wb.create_sheet("Sources")

    headers = [
        "Source Name",
        "Source Type",
        "Year / Period",
        "Used For",
        "Link / Reference",
        "Validation Notes",
    ]

    ws.append(headers)
    style_header(ws[1])

    sample_rows = [
        ["Millennium bcp Annual Report 2025", "Annual Report", "2025A", "Historical financials and ratios", "", "Pending"],
        ["Millennium bcp Annual Report 2024", "Annual Report", "2024A", "Historical financials and ratios", "", "Pending"],
        ["Millennium bcp Annual Report 2023", "Annual Report", "2023A", "Historical financials and ratios", "", "Pending"],
        ["Millennium bcp Annual Report 2022", "Annual Report", "2022A", "Historical financials and ratios", "", "Pending"],
        ["Euronext Lisbon", "Market Data", "Selected valuation date", "Share price and market data", "", "Pending"],
    ]

    for row in sample_rows:
        ws.append(row)

    autofit_columns(ws)


def create_historical_financials_sheet(wb):
    ws = wb.create_sheet("Historical Financials")

    headers = ["Metric", "Category"] + PERIODS[:4] + ["Unit", "Source Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    metrics = [
        ["Net interest income", "Income Statement"],
        ["Fees and commissions", "Income Statement"],
        ["Other operating income", "Income Statement"],
        ["Operating income", "Income Statement"],
        ["Operating costs", "Income Statement"],
        ["Impairments and provisions", "Income Statement"],
        ["Profit before tax", "Income Statement"],
        ["Net income", "Income Statement"],
        ["Loans to customers", "Balance Sheet"],
        ["Customer deposits", "Balance Sheet"],
        ["Total assets", "Balance Sheet"],
        ["Equity", "Balance Sheet"],
        ["Risk-weighted assets", "Capital"],
        ["CET1 capital", "Capital"],
        ["Number of shares", "Market / Per Share"],
        ["Share price", "Market Data"],
        ["Market capitalisation", "Market Data"],
    ]

    for metric, category in metrics:
        ws.append([metric, category, "", "", "", "", "EUR million", "Pending", "To be extracted from public sources"])

    autofit_columns(ws)


def create_ratios_sheet(wb):
    ws = wb.create_sheet("Ratios")

    headers = ["Ratio", "Category", "Formula"] + PERIODS[:4] + ["Unit", "Source Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    ratios = [
        ["ROE", "Profitability", "Net income / Average equity"],
        ["ROA", "Profitability", "Net income / Average total assets"],
        ["Net interest margin", "Profitability", "Net interest income / Average interest-earning assets"],
        ["Cost-to-income", "Efficiency", "Operating costs / Operating income"],
        ["Cost of risk", "Asset Quality", "Loan impairments / Average loans to customers"],
        ["NPL / NPE ratio", "Asset Quality", "Non-performing loans or exposures / Total loans or exposures"],
        ["NPL / NPE coverage ratio", "Asset Quality", "Impairment allowances / Non-performing loans or exposures"],
        ["Loan-to-deposit ratio", "Liquidity / Funding", "Loans to customers / Customer deposits"],
        ["CET1 ratio", "Capital", "Common Equity Tier 1 capital / Risk-weighted assets"],
        ["Total capital ratio", "Capital", "Total capital / Risk-weighted assets"],
        ["EPS", "Per Share", "Net income attributable to shareholders / Average number of shares"],
        ["Book value per share", "Per Share", "Equity attributable to shareholders / Number of shares"],
        ["Price-to-book", "Valuation", "Market capitalisation / Equity attributable to shareholders"],
        ["Price-to-earnings", "Valuation", "Market capitalisation / Net income attributable to shareholders"],
        ["Dividend yield", "Valuation", "Dividend per share / Share price"],
        ["Payout ratio", "Capital Return", "Dividends paid / Net income attributable to shareholders"],
    ]

    for ratio, category, formula in ratios:
        unit = "x" if ratio in ["Price-to-book", "Price-to-earnings"] else "%"
        if ratio in ["EPS", "Book value per share"]:
            unit = "EUR"

        ws.append([ratio, category, formula, "", "", "", "", unit, "Pending", "To be completed after data extraction"])

    autofit_columns(ws)


def create_assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions")

    headers = ["Assumption", "Category", "Conservative", "Base", "Optimistic", "Unit", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    assumptions = [
        ["Net interest income growth", "Revenue", "%"],
        ["Fees and commissions growth", "Revenue", "%"],
        ["Other operating income growth", "Revenue", "%"],
        ["Operating cost growth", "Costs", "%"],
        ["Cost-to-income target", "Efficiency", "%"],
        ["Cost of risk", "Asset Quality", "%"],
        ["Loan growth", "Balance Sheet", "%"],
        ["Deposit growth", "Balance Sheet", "%"],
        ["Total asset growth", "Balance Sheet", "%"],
        ["Equity growth", "Capital", "%"],
        ["Dividend payout ratio", "Capital Return", "%"],
        ["CET1 ratio assumption", "Capital", "%"],
        ["ROE assumption", "Profitability", "%"],
        ["P/B multiple", "Valuation", "x"],
        ["P/E multiple", "Valuation", "x"],
    ]

    for assumption, category, unit in assumptions:
        ws.append([assumption, category, "", "", "", unit, "To be completed based on public information and scenario logic"])

    autofit_columns(ws)


def create_forecast_sheet(wb):
    ws = wb.create_sheet("Forecast")

    headers = ["Metric", "Category"] + PERIODS[3:] + ["Driver", "Scenario", "Unit", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    forecast_metrics = [
        ["Net interest income", "Income Statement", "Net interest income growth"],
        ["Fees and commissions", "Income Statement", "Fees and commissions growth"],
        ["Other operating income", "Income Statement", "Other operating income assumption"],
        ["Operating income", "Income Statement", "Sum of revenue lines"],
        ["Operating costs", "Income Statement", "Operating cost growth"],
        ["Impairments and provisions", "Income Statement", "Cost of risk assumption"],
        ["Profit before tax", "Income Statement", "Operating profit after impairments"],
        ["Net income", "Income Statement", "Net income margin / tax assumption"],
        ["Loans to customers", "Balance Sheet", "Loan growth"],
        ["Customer deposits", "Balance Sheet", "Deposit growth"],
        ["Total assets", "Balance Sheet", "Asset growth"],
        ["Equity", "Balance Sheet", "Retained earnings and dividend assumption"],
        ["ROE", "Profitability Ratio", "Net income / average equity"],
        ["ROA", "Profitability Ratio", "Net income / average total assets"],
        ["Cost-to-income", "Efficiency Ratio", "Operating costs / operating income"],
        ["Loan-to-deposit ratio", "Liquidity Ratio", "Loans to customers / customer deposits"],
        ["CET1 ratio", "Capital Ratio", "Capital assumption"],
    ]

    for metric, category, driver in forecast_metrics:
        unit = "%" if "Ratio" in category or metric in ["ROE", "ROA", "Cost-to-income", "CET1 ratio"] else "EUR million"
        ws.append([metric, category, "", "", "", "", driver, "Base", unit, "Forecast figures are illustrative"])

    autofit_columns(ws)


def create_scenarios_sheet(wb):
    ws = wb.create_sheet("Scenarios")

    headers = ["Scenario", "Description", "Main Use"]
    ws.append(headers)
    style_header(ws[1])

    rows = [
        ["Conservative", "Lower profitability, weaker revenue growth, higher cost pressure or higher cost of risk", "Downside scenario"],
        ["Base", "Moderate continuation of current trends based on public information", "Central educational scenario"],
        ["Optimistic", "Stronger profitability, better efficiency or lower cost of risk", "Upside scenario"],
    ]

    for row in rows:
        ws.append(row)

    autofit_columns(ws)


def create_valuation_sheet(wb):
    ws = wb.create_sheet("Valuation")

    headers = ["Valuation Method", "Input", "Conservative", "Base", "Optimistic", "Unit", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    rows = [
        ["Price-to-book", "Book value / equity", "", "", "", "x", "Illustrative banking valuation multiple"],
        ["Price-to-earnings", "Net income", "", "", "", "x", "Illustrative earnings-based multiple"],
        ["Dividend yield", "Dividend per share / share price", "", "", "", "%", "Only if dividend data is available"],
        ["Scenario valuation", "Scenario output", "", "", "", "EUR million", "Educational valuation range only"],
    ]

    for row in rows:
        ws.append(row)

    autofit_columns(ws)


def create_sensitivity_sheet(wb):
    ws = wb.create_sheet("Sensitivity")

    headers = ["Sensitivity Table", "Variable 1", "Variable 2", "Output", "Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    rows = [
        ["ROE vs P/B", "ROE", "P/B multiple", "Illustrative valuation", "Pending", "To be completed"],
        ["Earnings growth vs P/E", "Earnings growth", "P/E multiple", "Illustrative valuation", "Pending", "To be completed"],
        ["Cost-to-income vs Cost of risk", "Cost-to-income", "Cost of risk", "Net income impact", "Pending", "To be completed"],
        ["NII growth vs Cost growth", "Net interest income growth", "Operating cost growth", "Operating profit impact", "Pending", "To be completed"],
    ]

    for row in rows:
        ws.append(row)

    autofit_columns(ws)


def create_checks_sheet(wb):
    ws = wb.create_sheet("Checks")

    headers = ["Check", "Purpose", "Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    checks = [
        ["Source check", "Confirm all key figures are mapped to public sources"],
        ["Unit check", "Confirm EUR million, percentage and multiple consistency"],
        ["Period check", "Confirm correct year and reporting period"],
        ["Formula check", "Confirm formulas are consistent"],
        ["Ratio check", "Compare reported ratios with calculated ratios where possible"],
        ["Scenario check", "Confirm assumptions differ logically across scenarios"],
        ["Disclaimer check", "Confirm no investment advice language is used"],
        ["Human review check", "Confirm reviewer has checked outputs before publication"],
    ]

    for check, purpose in checks:
        ws.append([check, purpose, "Pending", ""])

    autofit_columns(ws)


def create_workbook():
    wb = Workbook()

    create_cover_sheet(wb)
    create_sources_sheet(wb)
    create_historical_financials_sheet(wb)
    create_ratios_sheet(wb)
    create_assumptions_sheet(wb)
    create_forecast_sheet(wb)
    create_scenarios_sheet(wb)
    create_valuation_sheet(wb)
    create_sensitivity_sheet(wb)
    create_checks_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"Workbook created: {OUTPUT_FILE}")


if __name__ == "__main__":
    create_workbook()

"""
Create Sensitivity Analysis Template – Millennium bcp Case Study

This script creates an Excel template for the sensitivity analysis used in the
Millennium bcp financial modeling case study.

The workbook is created for educational and portfolio purposes only and does not
constitute investment advice.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "sensitivity_analysis.xlsx"


def style_header(row):
    """Apply basic styling to a header row."""
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(ws):
    """Apply a simple column width adjustment."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3


def create_cover_sheet(wb):
    ws = wb.active
    ws.title = "Cover"

    rows = [
        ["Sensitivity Analysis – Millennium bcp Case Study"],
        [""],
        ["Purpose", "Educational sensitivity analysis for a simplified banking financial model"],
        ["Company", "Millennium bcp / Banco Comercial Português"],
        ["Author", "Ricardo Serôdio"],
        [""],
        [
            "Disclaimer",
            "This analysis is for educational and portfolio purposes only and does not constitute investment advice.",
        ],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    autofit_columns(ws)


def create_inputs_sheet(wb):
    ws = wb.create_sheet("Inputs")

    headers = ["Input", "Category", "Low", "Base", "High", "Unit", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    rows = [
        ["ROE", "Profitability", "", "", "", "%", "Key driver of bank valuation"],
        ["Cost of equity", "Valuation", "", "", "", "%", "Used for valuation sensitivity"],
        ["Earnings growth", "Profitability", "", "", "", "%", "Affects earnings-based valuation"],
        ["Net interest income growth", "Revenue", "", "", "", "%", "Key revenue driver for banks"],
        ["Fees and commissions growth", "Revenue", "", "", "", "%", "Non-interest income driver"],
        ["Operating cost growth", "Costs", "", "", "", "%", "Affects cost-to-income and profitability"],
        ["Cost-to-income", "Efficiency", "", "", "", "%", "Operating efficiency ratio"],
        ["Cost of risk", "Asset Quality", "", "", "", "%", "Credit risk and impairment assumption"],
        ["P/B multiple", "Valuation", "", "", "", "x", "Book-value-based valuation multiple"],
        ["P/E multiple", "Valuation", "", "", "", "x", "Earnings-based valuation multiple"],
        ["Dividend payout", "Capital Return", "", "", "", "%", "Use only if dividend assumptions are included"],
    ]

    for row in rows:
        ws.append(row)

    autofit_columns(ws)


def create_roe_pb_sheet(wb):
    ws = wb.create_sheet("ROE vs P-B")

    rows = [
        ["ROE vs Price-to-Book Sensitivity"],
        [""],
        ["ROE / P-B", "Low P/B", "Base P/B", "High P/B"],
        ["Low ROE", "Pending", "Pending", "Pending"],
        ["Base ROE", "Pending", "Pending", "Pending"],
        ["High ROE", "Pending", "Pending", "Pending"],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    style_header(ws[3])
    autofit_columns(ws)


def create_growth_pe_sheet(wb):
    ws = wb.create_sheet("Growth vs P-E")

    rows = [
        ["Earnings Growth vs Price-to-Earnings Sensitivity"],
        [""],
        ["Earnings Growth / P-E", "Low P/E", "Base P/E", "High P/E"],
        ["Low Growth", "Pending", "Pending", "Pending"],
        ["Base Growth", "Pending", "Pending", "Pending"],
        ["High Growth", "Pending", "Pending", "Pending"],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    style_header(ws[3])
    autofit_columns(ws)


def create_efficiency_risk_sheet(wb):
    ws = wb.create_sheet("Efficiency vs Risk")

    rows = [
        ["Cost-to-Income vs Cost of Risk Sensitivity"],
        [""],
        ["Cost-to-Income / Cost of Risk", "Low Cost of Risk", "Base Cost of Risk", "High Cost of Risk"],
        ["Low Cost-to-Income", "Pending", "Pending", "Pending"],
        ["Base Cost-to-Income", "Pending", "Pending", "Pending"],
        ["High Cost-to-Income", "Pending", "Pending", "Pending"],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    style_header(ws[3])
    autofit_columns(ws)


def create_nii_cost_sheet(wb):
    ws = wb.create_sheet("NII vs Costs")

    rows = [
        ["Net Interest Income Growth vs Operating Cost Growth"],
        [""],
        ["NII Growth / Cost Growth", "Low Cost Growth", "Base Cost Growth", "High Cost Growth"],
        ["Low NII Growth", "Pending", "Pending", "Pending"],
        ["Base NII Growth", "Pending", "Pending", "Pending"],
        ["High NII Growth", "Pending", "Pending", "Pending"],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    style_header(ws[3])
    autofit_columns(ws)


def create_output_summary_sheet(wb):
    ws = wb.create_sheet("Output Summary")

    headers = ["Sensitivity Area", "Key Variables", "Output", "Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    rows = [
        ["ROE vs P/B", "ROE and P/B multiple", "Illustrative valuation range", "Pending", "To be completed after model inputs"],
        ["Growth vs P/E", "Earnings growth and P/E multiple", "Illustrative earnings valuation range", "Pending", "To be completed after model inputs"],
        ["Efficiency vs Risk", "Cost-to-income and cost of risk", "Profitability impact", "Pending", "To be completed after model inputs"],
        ["NII vs Costs", "Net interest income growth and cost growth", "Operating profit impact", "Pending", "To be completed after model inputs"],
    ]

    for row in rows:
        ws.append(row)

    autofit_columns(ws)


def create_checks_sheet(wb):
    ws = wb.create_sheet("Checks")

    headers = ["Check", "Purpose", "Status", "Notes"]
    ws.append(headers)
    style_header(ws[1])

    checks = [
        ["Input check", "Confirm all sensitivity inputs are documented"],
        ["Formula check", "Confirm sensitivity tables calculate correctly"],
        ["Unit check", "Confirm percentages and multiples are formatted correctly"],
        ["Scenario check", "Confirm low, base and high assumptions are logical"],
        ["Output check", "Confirm outputs are shown as ranges, not recommendations"],
        ["Disclaimer check", "Confirm no investment advice language is used"],
        ["Human review check", "Confirm reviewer has checked outputs before publication"],
    ]

    for check, purpose in checks:
        ws.append([check, purpose, "Pending", ""])

    autofit_columns(ws)


def create_workbook():
    wb = Workbook()

    create_cover_sheet(wb)
    create_inputs_sheet(wb)
    create_roe_pb_sheet(wb)
    create_growth_pe_sheet(wb)
    create_efficiency_risk_sheet(wb)
    create_nii_cost_sheet(wb)
    create_output_summary_sheet(wb)
    create_checks_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"Workbook created: {OUTPUT_FILE}")


if __name__ == "__main__":
    create_workbook()
